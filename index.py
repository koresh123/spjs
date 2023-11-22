# -*- coding: utf-8 -*-
"""
Created on Fri Nov 10 15:11:29 2023

@author: ava.kornilov
"""

import os
import re
import requests
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from loguru import logger
from configparser import ConfigParser
from exchangelib import (Credentials, Account, Configuration,
                         DELEGATE, Message, FileAttachment)

config = ConfigParser()
config.read('config.ini')

username = config['credentials']['username']
password = config['credentials']['password']

BASE_LINK = config['http']['base_link_mgps']

LIST_NAME_CELL_IN_EXCEL_REZAULT = [
    'П№',
    'Номер заявления',
    'Дата изменения статуса',
    'Кредитор',
    'Признак Фрод',
    'Признак дубля по кадастровому',
    'ФИО ответственного',
    ]

LIST_POZITIV_STATUS_DUBL = [
    'Внесено в Реестр,включено в заявку на выплату',
    'Заявка на выплату исполнена',
    'Уведомление о погашении задолженности',
    'Поступление возврата средств',
    'Исполнено',
    'Исполнено (переплата)',
    'Заключение сформировано/Просмотр заключения',
    ]

LIST_STATUS_RUCHNAYA_PROVERKA = [
    'На доработке',
    'Подписано и отправлено',
    'На рассмотрении(Первичная оценка)',
    'На рассмотрении (Запрос выписок ЕГРН, ЕГР ЗАГС)',
    'На рассмотрении (Ожидание ответов от Банка, ЕГРН, ЕГР ЗАГС)',
    'На рассмотрении (Выписки ЕГРН, ЕГР ЗАГС получены)',
    'На рассмотрении (Ответы от Банка, ЕГРН, ЕГР ЗАГС получены)',
    'Сформировано',
    'Сформировано повторно',
    'Формирование заключения',
    ]

LIST_STATUS_MGPS = []
for status_mgps in config['status']:
    LIST_STATUS_MGPS.append(config['status'][status_mgps])

LIST_TO_MAIL = []
for mail in config['to_mail']:
    LIST_TO_MAIL.append(config['to_mail'][mail])

SUBJECT_MAIL = 'Отчет робота СПЖС5'


class FunctionsStatic:
    @staticmethod
    def unescape(link: str) -> str:
        link = link.replace("&lt;", "<")
        link = link.replace("&gt;", ">")
        link = link.replace("&amp;", "&")
        return link

    @staticmethod
    def is_difference_more_than_year(date_1: str, date_2: str) -> bool:
        date_1_obj = datetime.strptime(date_1, '%Y-%m-%dT%H:%M:%S')
        date_2_obj = datetime.strptime(date_2, '%Y-%m-%dT%H:%M:%S')
        difference_date = date_1_obj - date_2_obj

        if difference_date.days <= 365:
            return True
        else:
            return False

    @staticmethod
    def subtract_until_smaller(k_naznacheni: int, len_df: int) -> int:
        while k_naznacheni >= len_df:
            difference = k_naznacheni - len_df
            k_naznacheni -= len_df
        return difference

    @staticmethod
    def f_date_conversion(date: str) -> str:
        date = date.split('T')[0]
        date = date.split('-')
        date = f'{date[2]}.{date[1]}.{date[0]}'
        return date

    @staticmethod
    def report_message(file_path: str, to_mail: list, subject: str):
        credentials = Credentials(
            username=config['bot_mail']['username'],
            password=config['bot_mail']['password'])

        conf = Configuration(
            server=config['bot_mail']['server'],
            credentials=credentials)

        account = Account(
            primary_smtp_address=config['bot_mail']['email'],
            credentials=credentials,
            autodiscover=False,
            config=conf,
            access_type=DELEGATE)

        m = Message(
            account=account,
            folder=account.sent,

            to_recipients=to_mail,
            subject=subject,
        )

        if file_path:
            with open(file_path, 'rb') as f:
                attachment_content = f.read()
            file = FileAttachment(
                name=os.path.split(file_path)[-1],
                content=attachment_content
            )
            m.attach(file)
        m.send_and_save()

    @staticmethod
    def formatirovanie_in_excel_file(path: str) -> None:
        workbook = load_workbook(path)
        worksheet = workbook.active

        fill = openpyxl.styles.PatternFill(
            start_color='99CCFF', end_color='99CCFF', fill_type='solid'
            )

        worksheet['A1'].fill = fill
        worksheet['B1'].fill = fill
        worksheet['C1'].fill = fill
        worksheet['D1'].fill = fill
        worksheet['E1'].fill = fill
        worksheet['F1'].fill = fill
        worksheet['G1'].fill = fill

        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 18
        worksheet.column_dimensions['C'].width = 22
        worksheet.column_dimensions['D'].width = 33
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 35
        worksheet.column_dimensions['G'].width = 40

        border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
            )
        for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column
                ):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

        workbook.save(path)


class MGPs(FunctionsStatic):

    def __init__(self, base_link_mgps: str,
                 username_mgps: str, password_mgps: str):
        self.base_link_mgps = base_link_mgps
        self.username_mgps = username_mgps
        self.password_mgps = password_mgps
        self.session_mgps = None
        self.logger = logger
        self.logger.add(
            "logi.txt", format="{time} {level} {message}", level="INFO")

    @logger.catch
    def authorization(self):
        logger.info('Начинаю авторизацию на сайте МГПС')
        self.session_mgps = requests.Session()
        params = {
            'client_id': 'mms-app',
            'redirect_uri': f'{self.base_link_mgps}login',
            'response_type': 'code',
            'scope': 'read write',
            'state': '12AZf6'
            }
        fst = self.session_mgps.get('https://xn--80az8a.xn--d1aqf.xn--p1ai/'
                                    'idm/auth/realms/DOMRF/'
                                    'protocol/openid-connect/auth',
                                    params=params,
                                    verify=False
                                    )
        link = re.findall(r'action="[^"]*"', fst.text)[0]
        link = link.split('"')[1]
        link = FunctionsStatic.unescape(link)
        data = {
                'username': self.username_mgps,
                'password': self.password_mgps,
                }
        self.session_mgps.post(link, data=data, verify=False)
        self.session_mgps.get(self.base_link_mgps)
        logger.info('Авторизация пройдена')
        return self.session_mgps

    @logger.catch
    def return_applications(self, status: str) -> list:
        logger.info(f'Запускаю процесс получения заявок со статусом {status}')
        params = {
                'initiatorType.id': '0',
                'initiatorType.name': 'Все',
                'list_status_id': status,
                'page': 1,
                'size': 10000,
                'sorting.num': 'ASC',
                'status': '[object Object]',
                'status': '[object Object]'
                }

        notifications = self.session_mgps.get(
            f'{self.base_link_mgps}n2o/data/list',
            verify=False,
            params=params
            ).json()
        logger.info(f'Заявки со статусом {status} получил.')
        return notifications

    def info_investment_of_loan_obligation(
            self, notification_id: str) -> tuple:
        params = {
            'list_edit_applicationForm_id': notification_id,
                }
        debt_info = self.session_mgps.get(
            f'{self.base_link_mgps}n2o/data/list/{notification_id}/'
            f'edit/application/creditCommitmentView',
            params=params,
            verify=False).json()

        try:
            debt = debt_info['list'][0]['debt']
            contractDate = debt_info['list'][0]['contractDate']
            cadaster_number = debt_info['list'][0]['dkpCadastralNumber']
            tagret = debt_info['list'][0]['purpose']['name']
            transaction_form = debt_info['list'][0]['dealType']['name']
        except Exception:
            try:
                debt = debt_info['list'][0]['debt']
            except Exception:
                debt = None
            try:
                contractDate = debt_info['list'][0]['contractDate']
            except Exception:
                contractDate = None
            try:
                cadaster_number = debt_info['list'][0]['dkpCadastralNumber']
            except Exception:
                cadaster_number = None
            try:
                tagret = debt_info['list'][0]['purpose']['name']
            except Exception:
                tagret = 'Другое'
            try:
                transaction_form = debt_info['list'][0]['dealType']['name']
            except Exception:
                transaction_form = 'Другое'

        return debt, contractDate, cadaster_number, tagret, transaction_form

    @logger.catch
    def appoint_a_person_in_charge(self, notification_id: str, user_id: str,
                                   name: str, notification_num: str) -> None:
        logger.info(f'По заявке {notification_num} назначаю '
                    'ответственных - {name}')
        data = {
         "id": notification_id,
         "employee": {
             "id": user_id,
             "fio": name,
             "surname": name.split(' ')[0],
             "name": name.split(' ')[1],
             "patronymic": name.split(' ')[2]
             },
         "status": {"code": "003-114"},
         "isBankAuthor": False,
         "isEpguInitiator": True
                                   }
        self.session_mgps.post(
            f'{self.base_link_mgps}n2o/data/list/'
            f'{notification_id}/edit/application/'
            f'domrfChangeEmployeeForm/submit',
            json=data,
            verify=False
            )
        logger.info('Назначил')


if __name__ == '__main__':
    #  создаем экземпляро класса МГПС для работы с ним
    MGPS_obj = MGPs(BASE_LINK, username, password)
    #  производим авторизацию на сайте МГПС
    MGPS_obj.authorization()
    #  считываем данные с ответственными из файла эксель
    df_responsible = pd.read_excel('Ответственные.xlsx').dropna()
    #  считываем датайфрейм который будем использовать для проверки
    #  кадастррового номера
    df_info_cadaster_number = pd.read_excel('cadaste_info.xlsx')
    df_info_cadaster_number = df_info_cadaster_number[
        ['cadastral_number', 'status']]
    #  k_number - поррядковый номер заявки
    k_number = 0
    #  n_assigned_applications - заявок передано автоматически
    n_assigned_applications = 0
    #  создаем датафрейм, куда будем записывать результаты проверок
    df = pd.DataFrame(columns=LIST_NAME_CELL_IN_EXCEL_REZAULT)
    #  начинаем перебор заявок. Для этого идем по каждому статусу
    for status_requests in LIST_STATUS_MGPS:
        #  возвращаем все заявки в данном статусу
        notifications = MGPS_obj.return_applications(status_requests)
        #  перебираем заявки для данного статуса
        for notification in notifications['list']:
            #  порядковый номер заявки в списке для данного статуса
            k_number += 1
            #  опрределяем номер заявки
            number_notification = notification['num']
            #  определяем дату изменения статуса
            statusChangeDate = notification['statusChangeDate']
            # преобразовываем дату, в вид %d.%m.%Y
            statusChangeDate = FunctionsStatic.f_date_conversion(
                statusChangeDate)
            #  определяем кредитора, если ошибка возвращаем как None
            try:
                creditor = notification['creditor']['name']
            except Exception:
                creditor = None
            #  Переходим к проверке Фрод
            #  Для выполнения условия Фрод, на вкладке кредитное обязательство
            #  первоначальная сумма кредита, руб. < 600к руб. и
            #  Дата получения/Дата формирования должны быть меньше
            #  чем один год назад
            Flag_Frod = True
            formationDate = notification['formationDate']
            borrowerApplicationReceiveDate = notification[
                'borrowerApplicationReceiveDate']
            #  возвращщаем со вкладки кредитные обязательства
            #  сумму кедита и дату заключения КД
            debt, contractDate, cadastr_number, tagret, form_sdelka =\
                MGPS_obj.info_investment_of_loan_obligation(notification['id'])
            #  Проверряем, выполняется ли первое условие для ФРОТ
            #  а именно с сайта вернулась не сумма None, а также
            #  данная сумма меньше 600к
            if debt is not None:
                if debt > 600_000:
                    Flag_Frod = False
            else:
                Flag_Frod = False
            #  Если первый признак для проверки ФРОТ выполнился
            #  переходим к проверки второго признака связанного с датой.
            if Flag_Frod:
                #  проверяем, кредит ли это или нет
                #  кредитные деньги, если заявка начинается на "e"
                if number_notification.startswith('e'):
                    if not FunctionsStatic.is_difference_more_than_year(
                            borrowerApplicationReceiveDate, contractDate):
                        Flag_Frod = False
                else:
                    if not FunctionsStatic.is_difference_more_than_year(
                            formationDate, contractDate):
                        Flag_Frod = False
            #  Проверяем на цель назначения
            if Flag_Frod:
                if tagret == 'Приобретение земельного участка для целей индивидуального жилищного строительства' or tagret == 'Приобретение готового жилого помещения':
                    pass
                else:
                    Flag_Frod = False
            #  Проверяем на форму сделки
            if Flag_Frod:
                if form_sdelka.lower() == 'дкп':
                    pass
                else:
                    Flag_Frod = False
            #  проверяем, прошел ли ФРОД все проверки.
            #  они пройдены в том случае, если Flag_Frod = True
            if Flag_Frod:
                Frod_info = 'Да'
            else:
                Frod_info = 'Нет'

            if cadastr_number:
                #  осуществляем поверку на кадастер
                #  1. Фильтруем по кадастровому номеру для данной заявки
                df_cadaster_filter = df_info_cadaster_number[
                    df_info_cadaster_number[
                        'cadastral_number'] == cadastr_number]
                #  получаем уникальные значения по заявке
                unique_values_previous_status = df_cadaster_filter[
                    'status'].unique().tolist()
                if len(unique_values_previous_status) > 0:
                    Flag_POZITIV_STATUS_DUBL = False
                    for unique_values in unique_values_previous_status:
                        if unique_values in LIST_STATUS_RUCHNAYA_PROVERKA:
                            Flag_POZITIV_STATUS_DUBL = True
                            info_dubl_cadastr = 'Да'
                            break
                    Flag_STATUS_RUCHNAYA_PROVERKA = False
                    if not Flag_POZITIV_STATUS_DUBL:
                        for unique_values in unique_values_previous_status:
                            if unique_values in LIST_STATUS_RUCHNAYA_PROVERKA:
                                Flag_STATUS_RUCHNAYA_PROVERKA = True
                                info_dubl_cadastr = 'Требуется ручная проверка'
                                break
                    if not Flag_POZITIV_STATUS_DUBL and not Flag_STATUS_RUCHNAYA_PROVERKA:
                        info_dubl_cadastr = 'Нет'
                    #  осуществляем проверку
                else:
                    info_dubl_cadastr = 'Требуется ручная проверка'
            else:
                info_dubl_cadastr = 'Нет'

            #  определяем ответсвенного. Если ответсвенного нет, назначаем его
            user_FIO = notification['reviewerInitials']
            if not user_FIO:
                if n_assigned_applications >= df_responsible.shape[0]:
                    number_stroka = FunctionsStatic.subtract_until_smaller(
                        n_assigned_applications, df_responsible.shape[0])
                else:
                    number_stroka = n_assigned_applications
                user_FIO = df_responsible.loc[number_stroka, 'ФИО']
                MGPS_ID = df_responsible.loc[number_stroka, 'MGPS_ID']
                # MGPS_obj.appoint_a_person_in_charge(
                #     notification['id'], MGPS_ID, FIO, notification['num'])
                n_assigned_applications += 1

            dict_info_requests = {
                'П№': k_number,
                'Номер заявления': notification['num'],
                'Дата изменения статуса': statusChangeDate,
                'Кредитор': creditor,
                'Признак Фрод': Frod_info,
                'Признак дубля по кадастровому': info_dubl_cadastr,
                'ФИО ответственного': user_FIO,
                }
            df = df.append(dict_info_requests, ignore_index=True)
    #  сохраняю результиррующий датафрейм в файл эксель
    df.to_excel('out.xlsx', index=False)
    #  применяю библиотеку openpyxl для визуального форматировани файла
    FunctionsStatic.formatirovanie_in_excel_file('out.xlsx')
    #  высылаю пписьмо с файлом о работе робота
    FunctionsStatic.report_message('out.xlsx', LIST_TO_MAIL, SUBJECT_MAIL)
